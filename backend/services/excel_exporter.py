from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime


NAVY = "1E3A5F"
WHITE = "FFFFFF"
BLUE_LIGHT = "DBEAFE"
GREEN_LIGHT = "DCFCE7"
YELLOW_LIGHT = "FEF9C3"
ORANGE_LIGHT = "FFEDD5"
GRAY_LIGHT = "F8FAFC"
ORANGE_DARK = "EA580C"


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _cell_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _cop(value) -> str:
    """Format number as Colombian pesos."""
    try:
        v = float(value)
        return f"${v:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(value)


def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def export_to_excel(filepath: str, project, material, machine, finishes, apu):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    results = apu.results_json.get("results", [])
    quantities = apu.quantities_json

    _build_sheet1(wb, project, material, machine, finishes, apu)
    _build_sheet2(wb, project, material, machine, finishes, apu, results, quantities)
    _build_sheet3(wb, project, results, quantities)

    wb.save(filepath)


def _build_sheet1(wb, project, material, machine, finishes, apu):
    ws = wb.create_sheet("Encabezado del Proyecto")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    def add_title_row(text, row, merge_cols=2):
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = _header_font(size=14)
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 30

    def add_row(label, value, row, label_fill=None, value_fill=None):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.font = _cell_font(bold=True)
        vc.font = _cell_font()
        lc.border = _border()
        vc.border = _border()
        if label_fill:
            lc.fill = _fill(label_fill)
        if value_fill:
            vc.fill = _fill(value_fill)

    r = 1
    add_title_row("ANÁLISIS DE PRECIOS UNITARIOS – EMPAQUES DE CARTÓN", r)
    r += 1
    add_title_row("DIFORMA", r)
    r += 1
    ws.row_dimensions[r].height = 10
    r += 1

    data_rows = [
        ("Código del Proyecto", project.code),
        ("Nombre del Proyecto", project.name),
        ("Cliente", project.client),
        ("Fecha", project.date_created.strftime("%d/%m/%Y") if project.date_created else ""),
        ("Estado", project.status.upper()),
        ("", ""),
        ("Material seleccionado", material.name),
        ("Tipo de material", material.material_type.capitalize()),
        ("Flauta / Calibre", material.flauta or material.calibre or "N/A"),
        ("Dirección de flauta", apu.flute_direction),
        ("Dimensiones desarrollo (cm)", f"{apu.developed_width_cm} × {apu.developed_height_cm}"),
        ("Pliego estándar (cm)", f"{material.sheet_width_cm} × {material.sheet_height_cm}"),
        ("", ""),
        ("Proceso / Máquina", machine.name),
        ("Tipo de proceso", machine.machine_type.replace("_", " ").capitalize()),
        ("Acabados aplicados", ", ".join(f.name for f in finishes) if finishes else "Sin acabados"),
        ("", ""),
        ("Preparado por", ""),
        ("Firma", ""),
        ("Fecha de exportación", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    for label, value in data_rows:
        add_row(label, value, r, label_fill=GRAY_LIGHT)
        r += 1

    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color=WHITE)


def _build_sheet2(wb, project, material, machine, finishes, apu, results, quantities):
    ws = wb.create_sheet("APU Detallado")

    num_cols = len(quantities)
    col_offset = 2  # column B onwards for quantities

    # Title
    ws.merge_cells(f"A1:{get_column_letter(1 + num_cols)}1")
    c = ws.cell(row=1, column=1, value="ANÁLISIS DE PRECIOS UNITARIOS – DETALLE POR CANTIDAD")
    c.font = _header_font(size=13)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.cell(row=2, column=1, value=f"Proyecto: {project.code} – {project.name}")
    ws.cell(row=2, column=1).font = _cell_font(bold=True)
    ws.merge_cells(f"A2:{get_column_letter(1 + num_cols)}2")

    # Header row with quantities
    header_row = 3
    ws.cell(row=header_row, column=1, value="Concepto")
    ws.cell(row=header_row, column=1).font = _header_font()
    ws.cell(row=header_row, column=1).fill = _fill(NAVY)
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 38

    for i, qty in enumerate(quantities):
        col = col_offset + i
        c = ws.cell(row=header_row, column=col, value=f"{qty:,} uds".replace(",", "."))
        c.font = _header_font()
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="center")
        _set_col_width(ws, get_column_letter(col), 18)

    def add_section_header(row, text, fill_color):
        ws.merge_cells(f"A{row}:{get_column_letter(1 + num_cols)}{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = _cell_font(bold=True, size=10)
        c.fill = _fill(fill_color)
        c.border = _border()

    def add_data_row(row, label, keys, fmt="cop", fill_color=None, bold=False):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _cell_font(bold=bold)
        lc.border = _border()
        if fill_color:
            lc.fill = _fill(fill_color)
        for i, res in enumerate(results):
            col = col_offset + i
            val = res.get(keys) if isinstance(keys, str) else res.get(keys[0])
            vc = ws.cell(row=row, column=col)
            if fmt == "cop":
                vc.value = float(val) if val is not None else 0
                vc.number_format = '#,##0'
            elif fmt == "int":
                vc.value = int(val) if val is not None else 0
            elif fmt == "float4":
                vc.value = float(val) if val is not None else 0
                vc.number_format = '#,##0.0000'
            vc.font = _cell_font(bold=bold)
            vc.border = _border()
            vc.alignment = Alignment(horizontal="right")
            if fill_color:
                vc.fill = _fill(fill_color)

    r = 4

    # Section: Material params
    add_section_header(r, "PARÁMETROS DE CONSUMO DE MATERIAL", GREEN_LIGHT)
    r += 1
    add_data_row(r, "Piezas por pliego", "pieces_per_sheet", fmt="int", fill_color=GREEN_LIGHT)
    r += 1
    add_data_row(r, "Hojas requeridas", "sheets_needed", fmt="int", fill_color=GREEN_LIGHT)
    r += 1
    add_data_row(r, "Área neta por pieza (m²)", "area_neta_m2", fmt="float4", fill_color=GREEN_LIGHT)
    r += 1
    add_data_row(r, "Área bruta total (m²)", "area_bruta_total_m2", fmt="float4", fill_color=GREEN_LIGHT)
    r += 1

    # Section: Lot costs
    add_section_header(r, "COSTOS DEL LOTE (COP)", BLUE_LIGHT)
    r += 1
    add_data_row(r, "Costo material", "material_cost", fill_color=BLUE_LIGHT)
    r += 1
    add_data_row(r, "Costo acabados", "finish_cost", fill_color=BLUE_LIGHT)
    r += 1
    add_data_row(r, "Costo troquel", "die_cost_total", fill_color=BLUE_LIGHT)
    r += 1
    add_data_row(r, "Setup máquina", "setup_machine_cost", fill_color=BLUE_LIGHT)
    r += 1
    add_data_row(r, "Mano de obra directa", "labor_cost_total", fill_color=BLUE_LIGHT)
    r += 1
    add_data_row(r, "Costo maquinaria", "machine_cost_total", fill_color=BLUE_LIGHT)
    r += 1
    add_data_row(r, "Insumos y consumibles", "supplies_cost", fill_color=BLUE_LIGHT)
    r += 1
    add_data_row(r, "Control calidad + empaque", "qc_cost", fill_color=BLUE_LIGHT)
    r += 1
    add_data_row(r, "COSTO DIRECTO TOTAL", "direct_cost_total", fill_color=BLUE_LIGHT, bold=True)
    r += 1

    # Section: Unit costs
    add_section_header(r, "COSTOS UNITARIOS (COP/unidad)", YELLOW_LIGHT)
    r += 1
    add_data_row(r, "Costo directo unitario", "direct_cost_unit", fill_color=YELLOW_LIGHT)
    r += 1
    add_data_row(r, "CIF + Gastos administrativos", "cif_admin_unit", fill_color=YELLOW_LIGHT)
    r += 1
    add_data_row(r, "Utilidad", "profit_unit", fill_color=YELLOW_LIGHT)
    r += 1
    add_data_row(r, "PRECIO NETO SIN IVA", "net_price_unit", fill_color=YELLOW_LIGHT, bold=True)
    r += 1
    add_data_row(r, "IVA 19%", "vat_amount_unit", fill_color=YELLOW_LIGHT)
    r += 1
    add_data_row(r, "PRECIO FINAL CON IVA", "final_price_unit", fill_color=YELLOW_LIGHT, bold=True)
    r += 1

    # Section: Lot totals
    add_section_header(r, "TOTALES DEL LOTE (COP)", ORANGE_LIGHT)
    r += 1
    add_data_row(r, "Total neto sin IVA", "net_price_lot", fill_color=ORANGE_LIGHT)
    r += 1
    add_data_row(r, "TOTAL CON IVA", "final_price_lot", fill_color=ORANGE_LIGHT, bold=True)
    r += 1

    # Technical notes
    r += 1
    ws.merge_cells(f"A{r}:{get_column_letter(1 + num_cols)}{r}")
    ws.cell(row=r, column=1, value="NOTAS TÉCNICAS:")
    ws.cell(row=r, column=1).font = _cell_font(bold=True)
    r += 1

    notes = _build_technical_notes(material, machine, finishes)
    for note in notes:
        ws.merge_cells(f"A{r}:{get_column_letter(1 + num_cols)}{r}")
        ws.cell(row=r, column=1, value=note)
        ws.cell(row=r, column=1).font = _cell_font()
        r += 1


def _build_sheet3(wb, project, results, quantities):
    ws = wb.create_sheet("Resumen Comparativo")

    num_cols = len(quantities)

    ws.merge_cells(f"A1:{get_column_letter(2 + num_cols)}1")
    c = ws.cell(row=1, column=1, value="RESUMEN COMPARATIVO DE CANTIDADES")
    c.font = _header_font(size=13)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28
    ws.column_dimensions["A"].width = 35

    headers = ["Concepto"] + [f"{q:,} uds".replace(",", ".") for q in quantities]
    for j, h in enumerate(headers):
        c = ws.cell(row=2, column=j + 1, value=h)
        c.font = _header_font()
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="center")
        _set_col_width(ws, get_column_letter(j + 1), 18)

    rows_data = [
        ("Costo unitario total", "direct_cost_unit"),
        ("Precio neto sin IVA (unit.)", "net_price_unit"),
        ("Precio final con IVA (unit.)", "final_price_unit"),
        ("Total lote sin IVA", "net_price_lot"),
        ("Total lote con IVA", "final_price_lot"),
        ("Ahorro vs primera cantidad (%)", "_savings"),
    ]

    # Find min cost unit column for highlighting
    min_unit_price = min(r["final_price_unit"] for r in results) if results else 0
    min_col = None
    for i, r in enumerate(results):
        if r["final_price_unit"] == min_unit_price:
            min_col = i

    for row_i, (label, key) in enumerate(rows_data):
        r = row_i + 3
        ws.cell(row=r, column=1, value=label).font = _cell_font(bold=True)
        for i, res in enumerate(results):
            col = i + 2
            if key == "_savings":
                if i == 0:
                    val = "–"
                    ws.cell(row=r, column=col, value=val)
                else:
                    base = results[0]["final_price_unit"]
                    curr = res["final_price_unit"]
                    pct = ((base - curr) / base * 100) if base > 0 else 0
                    c = ws.cell(row=r, column=col, value=f"{pct:.1f}%")
                    c.alignment = Alignment(horizontal="center")
            else:
                c = ws.cell(row=r, column=col, value=float(res.get(key, 0)))
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal="right")
                # Highlight best column
                if i == min_col:
                    c.fill = _fill("BBF7D0")
                    c.font = _cell_font(bold=True)
            ws.cell(row=r, column=col).border = _border()
        ws.cell(row=r, column=1).border = _border()

    # Recommendation text
    rec_row = len(rows_data) + 4
    best_qty = quantities[min_col] if min_col is not None else quantities[0]
    best_price = results[min_col]["final_price_unit"] if min_col is not None else 0
    ws.merge_cells(f"A{rec_row}:{get_column_letter(1 + num_cols)}{rec_row}")
    msg = f"Cantidad recomendada: {best_qty:,} unidades – Costo unitario más eficiente: {_cop(best_price)}"
    c = ws.cell(row=rec_row, column=1, value=msg.replace(",", "."))
    c.font = Font(name="Calibri", size=11, bold=True, color=ORANGE_DARK)
    c.alignment = Alignment(horizontal="center")

    # Bar chart
    if len(results) > 1:
        try:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Precio unitario con IVA por cantidad"
            chart.y_axis.title = "COP"
            chart.x_axis.title = "Cantidad"

            data_row_start = 4  # "Costo unitario total" row
            data_ref = Reference(ws, min_col=2, max_col=1 + num_cols, min_row=data_row_start, max_row=data_row_start)
            cats = Reference(ws, min_col=2, max_col=1 + num_cols, min_row=2, max_row=2)
            chart.add_data(data_ref)
            chart.set_categories(cats)
            chart.shape = 4
            ws.add_chart(chart, f"A{rec_row + 2}")
        except Exception:
            pass


def _build_technical_notes(material, machine, finishes):
    notes = []
    finish_names = [f.name.lower() for f in finishes]

    is_heavy = material.flauta in ("BC", "AC") if material.flauta else False
    has_laminate = any("laminad" in n for n in finish_names)

    if is_heavy and has_laminate:
        notes.append(
            "⚠ Nota: Por condiciones climáticas de Bogotá (alta altitud y humedad), "
            "se recomienda un tiempo de curado de 12 horas en estiba entre el colaminado y el troquelado."
        )

    if machine.machine_type == "cnc":
        notes.append(
            "ℹ Nota: La mesa CNC no requiere troquel físico. El costo de troquel no aplica para este proceso."
        )

    if material.material_type == "plegadizo":
        has_offset = any("cmyk" in n or "offset" in n or "4 col" in n for n in finish_names)
        if has_offset:
            notes.append(
                "ℹ Nota: El pliego estándar Maule de 70×100 cm requiere prensa offset de formato estándar."
            )

    if not notes:
        notes.append("Sin notas técnicas adicionales para esta combinación de material y proceso.")

    return notes
